import openpyxl
import sys
import time
Lcol=[2,3,4,5,6,7,8,9,10,11]
workbook = openpyxl.load_workbook("/usr/local/tomcat9/webapps/ExperimentReport/scripts/MeasurementOfSound.xlsx")
sheet = workbook.get_sheet_by_name("Sheet1")
count=1
for col in Lcol:
    sheet.cell(row=3,column=col).value=float(sys.argv[count])
    count+=1
for col in Lcol:
    sheet.cell(row=5,column=col).value=float(sys.argv[count])
    count+=1
sheet.cell(row=8,column=7).value=float(sys.argv[count])
count+=1
sheet.cell(row=9,column=7).value=float(sys.argv[count])
t=time.time()
workbook.save("/usr/local/tomcat9/webapps/ExperimentReport/downloads/MeasurementOfSound" + str(t) + ".xlsx")
print(",,"+str(t))


