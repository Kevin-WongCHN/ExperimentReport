
import openpyxl
import sys
import time

workbook = openpyxl.load_workbook("/usr/local/tomcat9/webapps/ExperimentReport/scripts/MillikanOilDrop.xlsx")
worksheet = workbook.get_sheet_by_name("Sheet1")
Ucol = [3, 5, 7, 9, 11]
tcol = [4, 6, 8, 10, 12]

arg = 1
for col in Ucol:
    for row in range(4, 9):
        worksheet.cell(row=row, column=col).value = float(sys.argv[arg])
        arg += 1

for col in tcol:
    for row in range(4, 9):
        worksheet.cell(row=row, column=col).value = float(sys.argv[arg])
        arg += 1

t=time.time()
workbook.save("/usr/local/tomcat9/webapps/ExperimentReport/downloads/MillikanOilDrop" + str(t) + ".xlsx")
print(",,"+str(t))