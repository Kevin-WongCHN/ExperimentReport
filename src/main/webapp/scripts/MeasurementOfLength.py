import openpyxl
import sys
import time

workbook = openpyxl.load_workbook("/usr/local/tomcat9/webapps/ExperimentReport/scripts/MeasurementOfLength.xlsx")
sheet = workbook.get_sheet_by_name("Sheet1")
count=1
rows=[3,4,5,6,7]
for row in rows:
    sheet.cell(row=row,column=2).value=float(sys.argv[count])
    count+=1
for row in rows:
    sheet.cell(row=row,column=3).value=float(sys.argv[count])
    count+=1
for row in rows:
    sheet.cell(row=row,column=4).value=float(sys.argv[count])
    count+=1
for row in rows:
    sheet.cell(row=row,column=5).value=float(sys.argv[count])
    count+=1
sheet.cell(row=3,column=8).value=float(sys.argv[count])
count+=1
sheet.cell(row=6,column=8).value=float(sys.argv[count])


t=time.time()
workbook.save("/usr/local/tomcat9/webapps/ExperimentReport/downloads/MeasurementOfLength" + str(t) + ".xlsx")
print(",,"+str(t))